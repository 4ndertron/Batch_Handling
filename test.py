import snowflake.connector
import os
import json
import openpyxl as xl


class BatchHandler:
    """
    This class is setup to do the following tasks:
        1) Create an excel workbook populated by two SQL queries.
        2) Run a headless chromedriver to upload the excel workbook to ShareFile.
    """
    new_account_sql_file = 'C:\\Users\\robert.anderson\\PycharmProjects\\' \
                           'UrgentFeedback\\SQL\\Filings\\FilingsNewAccounts.sql'
    transfer_accounts_sql_file = 'C:\\Users\\robert.anderson\\PycharmProjects\\' \
                                 'UrgentFeedback\\SQL\\Filings\\SepTransferCases.sql'
    dl_dir = os.path.join(os.environ['USERPROFILE'], 'Downloads')

    def __init__(self):
        self.workbook = None
        self.sheet_1 = None
        self.sheet_2 = None
        self.user = ''
        self.password = ''
        self.account = ''
        self.warehouse = ''
        self.database = ''
        self.new_account_sql = ''
        self.transfer_accounts_sql = ''

    def _set_credentials(self):
        snowflake_json = json.loads(os.environ['SNOWFLAKE_KEY'])
        self.user = snowflake_json['USERNAME']
        self.password = snowflake_json['PASSWORD']
        self.account = snowflake_json['ACCOUNT']
        self.warehouse = snowflake_json['WAREHOUSE']
        self.database = snowflake_json['DATABASE']

    def _collect_queries(self):
        open_sql = open(self.new_account_sql_file, 'r')
        self.new_account_sql = open_sql.read()
        open_sql.close()
        open_sql = open(self.transfer_accounts_sql_file, 'r')
        self.transfer_accounts_sql = open_sql.read()
        open_sql.close()

    def _create_workbook(self):
        """
        This function is designed to save memory space until it is needed to generate the actual workbook.
        :return:
        """
        self.workbook = xl.Workbook('temp_batch.xlsx')
        self.sheet_1 = self.workbook.create_sheet('New Accounts', 0)
        self.sheet_2 = self.workbook.create_sheet('Transfer Accounts', 1)

    def _populate_workbook(self):
        con = snowflake.connector.connect(
            user=self.user,
            password=self.password,
            account=self.account,
            warehouse=self.warehouse,
            database=self.database
        )
        # TODO: Reduce the new account results to just the most recent batch before adding the values to the workbook.
        cur = con.cursor().execute(self.new_account_sql)
        col_names = [x[0] for x in cur.description]
        results = cur.fetchall()
        self.sheet_1.append(col_names)
        for result in results:
            self.sheet_1.append(result)
        cur = con.cursor().execute(self.transfer_accounts_sql)
        results = cur.fetchall()
        col_names = [x[0] for x in cur.description]
        self.sheet_2.append(col_names)
        for result in results:
            self.sheet_2.append(result)
        cur.close()
        con.close()
        self.workbook.save(os.path.join(self.dl_dir, 'Batch {}.xlsx'.format(str(292))))
        self.workbook.close()

    def run_batch_handler(self):
        self._set_credentials()
        self._collect_queries()
        self._create_workbook()
        self._populate_workbook()


if __name__ == '__main__':
    new_handle = BatchHandler()
    new_handle.run_batch_handler()
