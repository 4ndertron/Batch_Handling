from . import os
from . import re
from . import json
from . import Enum
from . import Validation
from . import project_dir
from . import temp_batch_dir
from .snowflake import SnowflakeHandler
import snowflake.connector  # todo: Replace the direct connector method with the local snowflake module.
import openpyxl as xl
import datetime as dt
import pandas as pd


class BatchHandler:
    """
    Prerequisites:
        1) You need to create a JSON formatted Environment Variable, named "SNOWFLAKE_KEY", with the following keys.
            1) USER
            2) PASSWORD
            3) ACCOUNT
            4) WAREHOUSE
            5) DATABASE
        Please see Snowflake documentation for the definitions of the required fields.
    """

    class Keys(Enum):
        """
        Class level keys used in kwargs
        """
        PRIMARY_TABLE = 'primary_table'
        SCHEMA = 'schema'
        FILE_STRUCTURE = 'file_structure'
        MISSING_FILE_TYPE = 'missing_file_type'
        FILE_NAME = 'File Name'

    class _DefaultParameters(Enum):
        """
        Class level parameters
        """
        SCHEMA = 'D_POST_INSTALL'
        PRIMARY_TABLE = 'T_FILING_BATCH_UPLOAD_FILES'
        BATCH_REGEX = re.compile(r'Batch \d{3}', re.I)
        MISSING_FILE_TYPE = Validation.missing_file_type_batch.value

    all_batch_sql_file = os.path.join(project_dir,
                                      'queries',
                                      'account_list.sql')

    def __init__(self, batch_end_date, save_location, console_output=False, *args, **kwargs):
        self.console_output = console_output
        self.schema = kwargs[self.Keys.SCHEMA.value] if self.Keys.SCHEMA.value in kwargs \
            else self._DefaultParameters.SCHEMA.value
        self.primary_table = kwargs[self.Keys.PRIMARY_TABLE.value] if self.Keys.PRIMARY_TABLE.value in kwargs \
            else self._DefaultParameters.PRIMARY_TABLE.value
        self.file_structure = kwargs[self.Keys.FILE_STRUCTURE.value] if self.Keys.FILE_STRUCTURE.value in kwargs \
            else Validation.weekly_batch_structure.value
        self.missing_file_type = kwargs[
            self.Keys.MISSING_FILE_TYPE.value] if self.Keys.MISSING_FILE_TYPE.value in kwargs \
            else self._DefaultParameters.MISSING_FILE_TYPE.value
        self.workbooks = {}
        self.sheets = []
        self.snowflake_credentials = {}
        self.args_passed = args
        self.kwargs_passed = kwargs
        self.snowflake = SnowflakeHandler(console_output=console_output,
                                          schema=self.schema,
                                          primary_table=self.primary_table)
        self.all_batch_sql = ''
        self.batch_number = 0
        self.save_location = save_location
        self.batch_end_date = batch_end_date
        self.con = None
        self.cur = None
        self.existing_batch_files_query = """SELECT DISTINCT F.FILE_NAME
FROM D_POST_INSTALL.T_FILING_BATCH_UPLOAD_FILES AS F
"""
        self.existing_invoice_files_query = """SELECT DISTINCT I.FILE_NAME
FROM D_POST_INSTALL.T_FILINGS_INVOICE_DETAILS AS I
"""

    def _set_credentials(self):
        if self.console_output:
            print('Collecting Snowflake credentials form system environment...')
        snowflake_json = json.loads(os.environ['SNOWFLAKE_KEY'])
        for k, v in snowflake_json.items():
            self.snowflake_credentials[k.lower()] = v
        if self.console_output:
            print('credentials have been collected and assigned')

    def _collect_queries(self):
        if self.console_output:
            print('Reading the queries...')
        open_sql = open(self.all_batch_sql_file, 'r')
        self.all_batch_sql = open_sql.read()
        open_sql.close()
        if self.console_output:
            print('I know what to ask Snowflake.')

    def _calc_batch(self):
        """
        This function will calculate the number of weeks passed since the start of the batch process.
        The batch process began on 2013-11-02
        :return: assigns self.batch_number the calculated batch number.
        """
        if self.console_output:
            print('Finding the batch number based on the batch date value passed during object creation...')
        start_date = dt.date(2013, 11, 2)
        diff = self.batch_end_date - start_date
        diff_weeks = diff.days / 7
        self.batch_number = int(diff_weeks)
        if self.console_output:
            print(f'Batch {self.batch_number} is being evaluated.')

    def _generate_workbooks(self):
        if self.console_output:
            print('Making a temporary workspace in memory')
        for file in self.file_structure:
            self.workbooks[file] = {}
            self.workbooks[file][Validation.workbook_key.value] = xl.Workbook(file + '.xlsx')
            self.workbooks[file][Validation.worksheet_key.value] = []
            for i in range(len(self.file_structure[file])):
                self.workbooks[file][Validation.worksheet_key.value].append(
                    self.workbooks[file][Validation.workbook_key.value].create_sheet(self.file_structure[file][i], i))

    def _populate_workbook(self):
        if self.console_output:
            print('Knocking on Snowflake\'s door...')
        self.con = snowflake.connector.connect(  # Create the snowflake connection in the class instance variable
            **self.snowflake_credentials  # Use the credentials found in the environment variable
        )
        if self.console_output:
            print('Asking Snowflake your question...')
        self.cur = self.con.cursor().execute(self.all_batch_sql)  # Retrieve the query results
        col_names = [x[0] for x in self.cur.description]  # Generate a list of column names
        if self.console_output:
            print('Collecting results...')
        results = self.cur.fetchall()  # Assign all the query records to a method variable
        if self.console_output:
            print('Writing results to the temporary workspace...')
        for wb in self.workbooks:
            for ws in self.workbooks[wb][Validation.worksheet_key.value]:
                ws.append(col_names)
        # for sheet in self.sheets:
        #     sheet.append(col_names)
        for result in results:  # Iterate through the results
            batch_index = len(result) - 2  # Get the index integer of the batch number column
            result_batch = result[batch_index]  # Get the batch value of the specific result
            if result_batch == self.batch_number:
                type_index = len(result) - 1  # Get the index integer of the batch type column
                result_type = result[type_index]  # Get the batch type value of the specific result
                for wb in self.workbooks:  # Iterate through the workbooks
                    if result_type in self.workbooks[wb][Validation.workbook_key.value].sheetnames:
                        self.workbooks[wb][Validation.worksheet_key.value][
                            self.workbooks[wb][Validation.workbook_key.value].sheetnames.index(result_type)].append(
                            result)  # Append the record to the sheet
        if self.console_output:
            print('Saving the workspaces to your designated location...')
        for wb in self.workbooks:
            dl_dir = Validation.dl_dir_match.value[self.save_location][wb]
            file_name = wb + f' {self.batch_number} {self.batch_end_date.strftime("%m.%d.%Y")}.xlsx'
            self.workbooks[wb][Validation.workbook_key.value].save(os.path.join(dl_dir, file_name))
            self.workbooks[wb][Validation.workbook_key.value].close()
        if self.console_output:
            print('All files have been saved, and all connections have been closed.')

    def _create_pd_df(self, full_file_path):
        xl = pd.ExcelFile(full_file_path)
        sheets = xl.sheet_names
        file = os.path.basename(full_file_path)
        file_name = Validation.file_name_re.value.match(file).group()
        file_df = pd.read_excel(full_file_path, sheets)
        compiled_df = None
        for page_df in file_df:
            if compiled_df is None:
                compiled_df = file_df[page_df]
            else:
                compiled_df = compiled_df.append(file_df[page_df])
        compiled_df[self.Keys.FILE_NAME.value] = file_name
        return compiled_df

    def _upload_missing_files(self):
        """
        This method compares the list of files in the ShareFile upload location to the list of files uploaded to the
        data warehouse. If there are files in ShareFile that are not in the data warehouse, this method will initiate
        the steps needed to convert the ShareFile .xlsx file to a .csv, and upload the new .csv file to the data
        warehouse table used in the initial search.
        :return: None
        :rtype: None
        """
        self.snowflake.set_con_and_cur()
        search_dir = Validation.file_type_search_dir.value[self.missing_file_type]
        current_files = os.listdir(search_dir)
        if self.missing_file_type == Validation.missing_file_type_batch.value:
            uploaded_files = [x[0] for x in self.snowflake.run_query_string(self.existing_batch_files_query)]
        elif self.missing_file_type == Validation.missing_file_type_invoice.value:
            uploaded_files = [x[0] for x in self.snowflake.run_query_string(self.existing_invoice_files_query)]
        else:  # lock the logic for the rest of the method to avoid any unintentional upload attempts.
            uploaded_files = current_files
        self.snowflake.close_con_and_cur()
        for file in current_files:
            file_mo = Validation.file_type_re.value[self.missing_file_type].match(file)
            file_name = Validation.file_name_re.value.match(file).group()
            if file_name not in uploaded_files and file_mo is not None:
                if self.console_output:
                    print(f'{file_name} needs to be uploaded')
                if not os.path.exists(temp_batch_dir):
                    os.mkdir(temp_batch_dir)
                temp_files = os.listdir(temp_batch_dir)
                if self.console_output:
                    print(f'project temp directory:\n{temp_files}')
                for temp in temp_files:
                    os.remove(os.path.join(temp_batch_dir, temp))
                if self.console_output:
                    print(f'project temp directory:\n{temp_files}')
                f_path = os.path.join(search_dir, file)
                compiled_df = self._create_pd_df(f_path)
                compiled_df.to_csv(os.path.join(temp_batch_dir, file_name + '.csv'), index=False)
                self.snowflake.run_table_updates(temp_batch_dir,
                                                 Validation.file_type_file_patterns.value[self.missing_file_type],
                                                 Validation.update_type_append.value)

    def upload_new_invoice(self, file_directory):
        """
        Full Method steps
        1) search email threads for a matching invoice detail thread from LD
        2) Download the invoice detail file
        3) Identify if the file name exists in the warehouse
        4) Convert the file to a csv, if it doesn't exist in the warehouse
        5) Upload the csv file to the warehouse, if one was made.
        6) Join the details of the new invoice with accounts sent in the batch files
        7) Identify the match ratio, and return findings to a text file to review the findings.
        :param file_directory:
        :type file_directory:
        :return:
        :rtype:
        """
        self.snowflake.set_con_and_cur()
        uploaded_files = [x[0] for x in self.snowflake.run_query_string(self.existing_invoice_files_query)]
        search_dir = file_directory
        current_files = os.listdir(search_dir)
        for file in current_files:
            file_name = Validation.file_name_re.value.match(file).group()
            if file_name not in uploaded_files:
                if self.console_output:
                    print(f'{file_name} needs to be uploaded to the warehouse')

    def run_batch_handler(self):
        self._set_credentials()
        self._collect_queries()
        self._generate_workbooks()
        self._calc_batch()
        self._populate_workbook()
        self._upload_missing_files()

    def run_new_invoices(self):
        self._set_credentials()
        self._upload_missing_files()

    def search_for_accounts(self, lst):
        """
        Pass a list of account numbers, as 6 or more digits, to have a dictionary returned of matching
        files in the default directory.
        NOTE: This method be obsolete since the batch files have been added to the data warehouse
        :param lst: list of 6 or more digit accounts .
        :return: dictionary of accounts found in a file.
        """
        batch_regex = re.compile(r'Batch \d{3,}.*', re.I)
        acct_regex = re.compile(r'\d{6,}')
        acct_list = {}
        sharefile_dir = Validation.dl_dir_match.value[Validation.batch_save_location_sharefile.value][
            Validation.Keys.value.BATCH.value]
        for file in os.listdir(sharefile_dir):
            if self.console_output:
                print('checking {}'.format(file))
            mo = batch_regex.search(file)
            if mo:
                if self.console_output:
                    print('opening {}'.format(file))
                f = xl.load_workbook(os.path.join(sharefile_dir, file))
                fs = f.active
                if self.console_output:
                    print('iterating through the first two columns...')
                for row in fs.iter_cols(min_col=1, max_col=2, min_row=2, max_row=fs.max_row):
                    rmo = acct_regex.search(str(row))
                    if rmo:
                        if int(rmo.string) in lst:
                            acct_list[rmo.string] = file
            else:
                if self.console_output:
                    print('{} did not match the batch regex'.format(file))
        if self.console_output:
            print('returning the dictionary and ending the function call.')
        return acct_list
